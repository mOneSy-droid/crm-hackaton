"""Mijozlar va sharhlarni Excel (.xlsx) ko'rinishida eksport qilish.

Maxfiylik qoidasi bu yerda ham amal qiladi: telefon raqami faqat
niqoblangan ko'rinishda (+9989****4567) chiqadi — egasiga ham to'liq
raqam berilmaydi, API bilan bir xil.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.logging import mask_phone
from app.models import Restaurant, Review, ReviewStatus

HEADER_FILL = PatternFill("solid", fgColor="0E1E45")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GUEST_NAME = "Mehmon"

STATUS_UZ = {
    ReviewStatus.PENDING: "Kutilmoqda",
    ReviewStatus.APPROVED: "Tasdiqlangan",
    ReviewStatus.REJECTED: "Rad etilgan",
}


def _style_header(sheet, widths: list[int]) -> None:
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
        cell = sheet.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22
    sheet.freeze_panes = "A2"


def _fmt_date(value: datetime | None) -> str:
    return value.strftime("%d.%m.%Y %H:%M") if value else ""


def build_customers_workbook(restaurant: Restaurant, reviews: list[Review]) -> BytesIO:
    """Ikki varaqli kitob: «Mijozlar» (jamlanma) va «Sharhlar» (to'liq ro'yxat)."""
    wb = Workbook()

    # -- 1-varaq: mijozlar jamlanmasi ---------------------------------------
    ws = wb.active
    ws.title = "Mijozlar"
    ws.append(["Mijoz", "Telegram", "Telefon", "Sharhlar soni", "O'rtacha baho", "Oxirgi sharh"])
    _style_header(ws, [28, 20, 18, 15, 15, 18])

    grouped: dict[object, list[Review]] = defaultdict(list)
    for review in reviews:
        grouped[review.author_id].append(review)

    for author_id, items in grouped.items():
        author = items[0].author
        name = (author.full_name if author and author.full_name else None) or GUEST_NAME
        username = f"@{author.telegram_username}" if author and author.telegram_username else ""
        phone = mask_phone(author.phone) if author else ""
        last = max((r.created_at for r in items), default=None)
        avg = sum(r.rating for r in items) / len(items)
        row = ws.max_row + 1
        ws.append([name, username, phone, len(items), round(avg, 2), _fmt_date(last)])
        ws.cell(row=row, column=5).number_format = "0.00"

    # -- 2-varaq: barcha sharhlar -------------------------------------------
    ws2 = wb.create_sheet("Sharhlar")
    ws2.append(["Sana", "Mijoz", "Baho", "Sharh matni", "Holat", "Javobingiz"])
    _style_header(ws2, [18, 26, 8, 60, 14, 45])

    for review in reviews:
        name = (
            review.author.full_name
            if review.author and review.author.full_name
            else GUEST_NAME
        )
        row = ws2.max_row + 1
        ws2.append(
            [
                _fmt_date(review.created_at),
                name,
                review.rating,
                review.text or "",
                STATUS_UZ.get(review.status, str(review.status)),
                review.owner_reply or "",
            ]
        )
        for col in (4, 6):
            ws2.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
